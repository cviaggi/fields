"""Spreadsheet utilities for data manipulation and file operations."""

import csv
import os
from typing import List, Dict, Any, Optional, Union, TextIO
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False


class Spreadsheet:
    """A class for handling spreadsheet-like data and CSV operations."""

    def __init__(self, data: Optional[List[Dict[str, Any]]] = None,
                 sheet_name: str = "Sheet1"):
        """Initialize the Spreadsheet with optional data.

        Args:
            data: List of dictionaries representing rows of data
            sheet_name: Name of the worksheet
        """
        # Initialize worksheets as a dictionary of sheet_name -> data
        self.worksheets = {}
        self.current_sheet = sheet_name

        if data is not None:
            self.worksheets[sheet_name] = data
        else:
            self.worksheets[sheet_name] = []

        self._update_current_data()

    def _update_current_data(self):
        """Update the current data reference based on current sheet."""
        self.data = self.worksheets[self.current_sheet]
        self.headers = self._extract_headers()

    def _extract_headers(self) -> List[str]:
        """Extract headers from the current worksheet data.

        Returns:
            List of column headers
        """
        if not self.data:
            return []

        # Get all unique keys from all rows
        headers = set()
        for row in self.data:
            headers.update(row.keys())

        # Sort headers for consistent ordering
        return sorted(list(headers))

    def create_sheet(self, sheet_name: str, data: Optional[List[Dict[str, Any]]] = None) -> None:
        """Create a new worksheet.

        Args:
            sheet_name: Name of the new worksheet
            data: Optional data for the new worksheet

        Raises:
            ValueError: If sheet name already exists
        """
        if sheet_name in self.worksheets:
            raise ValueError(f"Sheet '{sheet_name}' already exists")

        self.worksheets[sheet_name] = data or []
        self.current_sheet = sheet_name
        self._update_current_data()

    def switch_sheet(self, sheet_name: str) -> None:
        """Switch to a different worksheet.

        Args:
            sheet_name: Name of the worksheet to switch to

        Raises:
            ValueError: If sheet doesn't exist
        """
        if sheet_name not in self.worksheets:
            raise ValueError(f"Sheet '{sheet_name}' does not exist")

        self.current_sheet = sheet_name
        self._update_current_data()

    def get_sheet_names(self) -> List[str]:
        """Get list of all worksheet names.

        Returns:
            List of worksheet names
        """
        return list(self.worksheets.keys())

    def sheet_exists(self, sheet_name: str) -> bool:
        if sheet_name in self.worksheets:
            return True
        return False

    def get_sheet_data(self, sheet_name: str) -> List[Dict[str, Any]]:
        """Get data from a specific worksheet.

        Args:
            sheet_name: Name of the worksheet

        Returns:
            Data from the specified worksheet

        Raises:
            ValueError: If sheet doesn't exist
        """
        if sheet_name not in self.worksheets:
            raise ValueError(f"Sheet '{sheet_name}' does not exist")

        return self.worksheets[sheet_name]

    def remove_sheet(self, sheet_name: str) -> None:
        """Remove a worksheet.

        Args:
            sheet_name: Name of the worksheet to remove

        Raises:
            ValueError: If trying to remove the only sheet or if sheet doesn't exist
        """
        if sheet_name not in self.worksheets:
            raise ValueError(f"Sheet '{sheet_name}' does not exist")

        if len(self.worksheets) <= 1:
            raise ValueError("Cannot remove the only remaining worksheet")

        del self.worksheets[sheet_name]

        # If we removed the current sheet, switch to another one
        if self.current_sheet == sheet_name:
            self.current_sheet = list(self.worksheets.keys())[0]
            self._update_current_data()

    def _extract_headers(self) -> List[str]:
        """Extract headers from the data.

        Returns:
            List of column headers
        """
        if not self.data:
            return []

        # Get all unique keys from all rows
        headers = set()
        for row in self.data:
            headers.update(row.keys())

        # Sort headers for consistent ordering
        return sorted(list(headers))

    def add_row(self, row: Dict[str, Any]) -> None:
        """Add a new row to the spreadsheet.

        Args:
            row: Dictionary representing a row of data
        """
        self.data.append(row)
        # Update headers if new columns are introduced
        new_headers = set(row.keys()) - set(self.headers)
        if new_headers:
            self.headers.extend(sorted(list(new_headers)))

    def add_rows(self, rows: List[Dict[str, Any]]) -> None:
        """Add multiple rows to the spreadsheet.

        Args:
            rows: List of dictionaries representing rows of data
        """
        for row in rows:
            self.add_row(row)

    def get_row(self, index: int) -> Optional[Dict[str, Any]]:
        """Get a row by index.

        Args:
            index: Row index (0-based)

        Returns:
            Row data as dictionary, or None if index is out of range
        """
        if 0 <= index < len(self.data):
            return self.data[index]
        return None

    def get_column(self, column_name: str) -> List[Any]:
        """Get all values for a specific column.

        Args:
            column_name: Name of the column

        Returns:
            List of values in the column
        """
        return [row.get(column_name) for row in self.data]

    def filter_rows(self, condition: callable) -> 'Spreadsheet':
        """Filter rows based on a condition function.

        Args:
            condition: Function that takes a row dict and returns bool

        Returns:
            New Spreadsheet instance with filtered data
        """
        filtered_data = [row for row in self.data if condition(row)]
        return Spreadsheet(filtered_data)

    def sort_by_column(self, column_name: str, reverse: bool = False) -> None:
        """Sort the data by a specific column in-place.

        Args:
            column_name: Column to sort by
            reverse: Sort in descending order if True
        """
        def sort_key(row):
            value = row.get(column_name)
            # Handle None values by placing them at the end
            return (value is None, value)

        self.data.sort(key=sort_key, reverse=reverse)

    def to_csv_string(self, delimiter: str = ',') -> str:
        """Convert the spreadsheet data to CSV string format.

        Args:
            delimiter: CSV delimiter (default: comma)

        Returns:
            CSV formatted string
        """
        if not self.data:
            return ""

        import io
        output = io.StringIO()

        if self.headers:
            writer = csv.DictWriter(output, fieldnames=self.headers, delimiter=delimiter)
            writer.writeheader()
            writer.writerows(self.data)
        else:
            writer = csv.writer(output, delimiter=delimiter)
            writer.writerows(self.data)

        return output.getvalue()

    def save_to_csv(self, file_path: Union[str, Path], delimiter: str = ',') -> None:
        """Save the spreadsheet data to a CSV file.

        Args:
            file_path: Path to save the CSV file
            delimiter: CSV delimiter (default: comma)
        """
        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)

        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            if self.headers:
                writer = csv.DictWriter(f, fieldnames=self.headers, delimiter=delimiter)
                writer.writeheader()
                writer.writerows(self.data)
            else:
                writer = csv.writer(f, delimiter=delimiter)
                writer.writerows(self.data)

    @classmethod
    def from_csv_file(cls, file_path: Union[str, Path],
                     delimiter: str = ',') -> 'Spreadsheet':
        """Load spreadsheet data from a CSV file.

        Args:
            file_path: Path to the CSV file
            delimiter: CSV delimiter (default: comma)

        Returns:
            Spreadsheet instance with loaded data

        Raises:
            FileNotFoundError: If the file doesn't exist
            csv.Error: If there's an error parsing the CSV
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"CSV file not found: {file_path}")

        data = []
        with open(file_path, 'r', newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            for row in reader:
                data.append(row)

        return cls(data)

    def save_to_excel(self, file_path: Union[str, Path],
                     include_headers: bool = True,
                     header_style: Optional[Dict[str, Any]] = None) -> None:
        """Save all worksheets to an Excel file.

        Args:
            file_path: Path to save the Excel file
            include_headers: Whether to include column headers (default: True)
            header_style: Optional styling for headers (font, fill, etc.)

        Raises:
            ImportError: If openpyxl is not installed
        """
        if not EXCEL_SUPPORT:
            raise ImportError("Excel support requires openpyxl library. Install with: pip install openpyxl")

        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)

        # Create a new workbook
        workbook = Workbook()

        # Default header styling
        if header_style is None:
            header_style = {
                'font': Font(bold=True, color="FFFFFF"),
                'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            }

        # Remove default sheet if we have worksheets to add
        if self.worksheets:
            workbook.remove(workbook.active)

        # Create a worksheet for each sheet in our data
        for sheet_name, sheet_data in self.worksheets.items():
            worksheet = workbook.create_sheet(title=sheet_name)

            # Extract headers for this sheet
            headers = self._extract_headers_for_data(sheet_data)

            # Write headers if requested
            if include_headers and headers:
                for col_num, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=1, column=col_num, value=header)
                    if header_style:
                        # Apply header styling
                        if 'font' in header_style:
                            cell.font = header_style['font']
                        if 'fill' in header_style:
                            cell.fill = header_style['fill']

            # Write data rows
            start_row = 2 if (include_headers and headers) else 1

            for row_num, row_data in enumerate(sheet_data, start_row):
                for col_num, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    worksheet.cell(row=row_num, column=col_num, value=value)

            # Auto-adjust column widths for this sheet
            for col_num, header in enumerate(headers, 1):
                column_letter = worksheet.cell(row=1, column=col_num).column_letter
                max_length = len(str(header))

                for row_num in range(1, len(sheet_data) + start_row):
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))

                # Set column width with some padding
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # If no worksheets were created, create an empty default sheet
        if not self.worksheets:
            worksheet = workbook.create_sheet(title="Sheet1")

        # Save the workbook
        workbook.save(str(file_path))

    def _extract_headers_for_data(self, data: List[Dict[str, Any]]) -> List[str]:
        """Extract headers from specific worksheet data.

        Args:
            data: Worksheet data to extract headers from

        Returns:
            List of column headers
        """
        if not data:
            return []

        # Get all unique keys from all rows
        headers = set()
        for row in data:
            headers.update(row.keys())

        # Sort headers for consistent ordering
        return sorted(list(headers))

    @classmethod
    def from_excel_file(cls, file_path: Union[str, Path],
                       header_row: int = 1) -> 'Spreadsheet':
        """Load all worksheets from an Excel file.

        Args:
            file_path: Path to the Excel file
            header_row: Row number containing headers (1-based, default: 1)

        Returns:
            Spreadsheet instance with all worksheets loaded

        Raises:
            ImportError: If openpyxl is not installed
            FileNotFoundError: If the file doesn't exist
            ValueError: If the Excel file cannot be read
        """
        if not EXCEL_SUPPORT:
            raise ImportError("Excel support requires openpyxl library. Install with: pip install openpyxl")

        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        try:
            workbook = load_workbook(str(file_path), read_only=True, data_only=True)

            # Create instance with first sheet as default
            instance = cls()

            # Clear default empty sheet
            instance.worksheets.clear()

            # Load each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                data = []
                headers = []

                # Read headers from specified row
                header_row_idx = header_row - 1  # Convert to 0-based indexing
                for col in range(1, worksheet.max_column + 1):
                    header_value = worksheet.cell(row=header_row_idx + 1, column=col).value
                    headers.append(str(header_value) if header_value is not None else f"Column_{col}")

                # Read data rows
                for row in range(header_row_idx + 2, worksheet.max_row + 1):
                    row_data = {}
                    has_data = False

                    for col, header in enumerate(headers, 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value is not None:
                            has_data = True
                        row_data[header] = cell_value

                    # Only add rows that have at least some data
                    if has_data:
                        data.append(row_data)

                instance.worksheets[sheet_name] = data

            workbook.close()

            # Set first sheet as current if worksheets exist
            if instance.worksheets:
                instance.current_sheet = list(instance.worksheets.keys())[0]
                instance._update_current_data()
            else:
                # Create empty default sheet if no worksheets found
                instance.worksheets["Sheet1"] = []
                instance.current_sheet = "Sheet1"
                instance._update_current_data()

            return instance

        except Exception as e:
            raise ValueError(f"Error reading Excel file {file_path}: {e}")

    def create_excel_template(self, file_path: Union[str, Path],
                            template_config: Optional[Dict[str, Any]] = None) -> None:
        """Create an Excel template file with predefined structure.

        Args:
            file_path: Path to save the template
            template_config: Configuration for the template structure
        """
        if not EXCEL_SUPPORT:
            raise ImportError("Excel support requires openpyxl library. Install with: pip install openpyxl")

        # Default template configuration
        if template_config is None:
            template_config = {
                "sheet_name": "Permit Data",
                "columns": ["Field Name", "Permit Type", "Time Slots", "Date", "Status"],
                "sample_data": [
                    ["Shoreline North Field", "Athletic Field Use", "Sat 8:00 AM - 1:00 PM", "2025-12-06", "Available"],
                    ["Central Park Field", "Recreational Use", "Wed 12:00 PM - 5:00 PM", "2025-12-10", "Booked"]
                ]
            }

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = template_config["sheet_name"]

        # Write headers
        headers = template_config["columns"]
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Write sample data
        sample_data = template_config.get("sample_data", [])
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = worksheet.cell(row=1, column=col_num).column_letter
            worksheet.column_dimensions[column_letter].width = 20

        # Add instructions sheet
        instructions_sheet = workbook.create_sheet("Instructions")
        instructions_sheet['A1'] = "Field Permit Data Template"
        instructions_sheet['A1'].font = Font(size=14, bold=True)

        instructions = [
            "",
            "This template is used for managing field permit data.",
            "",
            "Columns:",
            "- Field Name: Name of the field/facility",
            "- Permit Type: Type of permit (Athletic, Recreational, etc.)",
            "- Time Slots: Available time slots for the field",
            "- Date: Date of availability/booking",
            "- Status: Current status (Available, Booked, Maintenance, etc.)",
            "",
            "Instructions:",
            "1. Fill in the data starting from row 2",
            "2. Add new rows as needed",
            "3. Save the file with permit data",
            "4. Import back into the Fields application"
        ]

        for row_num, instruction in enumerate(instructions, 2):
            instructions_sheet.cell(row=row_num, column=1, value=instruction)

        instructions_sheet.column_dimensions['A'].width = 60

        # Save the template
        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(file_path))

    @classmethod
    def from_csv_string(cls, csv_string: str, delimiter: str = ',') -> 'Spreadsheet':
        """Load spreadsheet data from a CSV string.

        Args:
            csv_string: CSV formatted string
            delimiter: CSV delimiter (default: comma)

        Returns:
            Spreadsheet instance with loaded data
        """
        import io
        data = []
        with io.StringIO(csv_string) as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            for row in reader:
                data.append(row)

        return cls(data)

    def get_summary_stats(self) -> Dict[str, Any]:
        """Get summary statistics for all worksheets in the spreadsheet.

        Returns:
            Dictionary with summary information for all worksheets
        """
        total_rows = sum(len(data) for data in self.worksheets.values())

        stats = {
            "total_sheets": len(self.worksheets),
            "total_rows": total_rows,
            "current_sheet": self.current_sheet,
            "sheet_stats": {}
        }

        # Calculate stats for each worksheet
        for sheet_name, sheet_data in self.worksheets.items():
            # Extract headers for this sheet
            headers = self._extract_headers_for_data(sheet_data)

            sheet_stats = {
                "rows": len(sheet_data),
                "columns": len(headers),
                "headers": headers.copy(),
                "column_stats": {}
            }

            # Calculate stats for each column in this sheet
            for header in headers:
                column_data = [row.get(header) for row in sheet_data]
                non_none_values = [v for v in column_data if v is not None]

                column_stats = {
                    "total_values": len(column_data),
                    "non_null_values": len(non_none_values),
                    "null_values": len(column_data) - len(non_none_values),
                    "unique_values": len(set(non_none_values))
                }

                # Try to detect data types
                if non_none_values:
                    types = set(type(v).__name__ for v in non_none_values)
                    column_stats["data_types"] = list(types)

                sheet_stats["column_stats"][header] = column_stats

            stats["sheet_stats"][sheet_name] = sheet_stats

        return stats

    def __len__(self) -> int:
        """Return the number of rows in the spreadsheet."""
        return len(self.data)

    def __getitem__(self, index: int) -> Dict[str, Any]:
        """Get a row by index."""
        return self.data[index]

    def __iter__(self):
        """Iterate over the rows."""
        return iter(self.data)

    def __repr__(self) -> str:
        """String representation of the spreadsheet."""
        total_rows = sum(len(data) for data in self.worksheets.values())
        total_sheets = len(self.worksheets)
        return f"Spreadsheet(sheets={total_sheets}, total_rows={total_rows}, current='{self.current_sheet}')"
